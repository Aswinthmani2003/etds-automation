from flask import Flask, request, jsonify
from openpyxl import load_workbook
from pyzbar.pyzbar import decode
from PIL import Image
import fitz
import re
import io

app = Flask(__name__)

def extract_digits(text):
    if text is None:
        return ""
    return re.sub(r'[^0-9]', '', str(text))

def parse_excel(excel_file):
    try:
        wb = load_workbook(excel_file)
        token_map = {}

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            rows = list(sheet.iter_rows(values_only=True))

            for row in rows:
                if not row:
                    continue

                # Look for 20-digit and 15-digit numbers in each row
                token = None
                target = None

                for cell in row:
                    d = extract_digits(cell)
                    if len(d) == 20 and not token:
                        token = d
                    if len(d) == 15 and not target:
                        target = d

                if token and target and token not in token_map:
                    token_map[token] = target

        return token_map
    except Exception as e:
        return None

def decode_barcode(pdf_bytes):
    try:
        pdf = fitz.open(stream=pdf_bytes, filetype="pdf")
        max_pages = min(pdf.page_count, 3)

        for page_num in range(max_pages):
            page = pdf[page_num]
            pix = page.get_pixmap(matrix=fitz.Matrix(2, 2))
            img = Image.frombytes("RGB", [pix.width, pix.height], pix.samples)

            decoded = decode(img)
            if decoded:
                for obj in decoded:
                    code = obj.data.decode('utf-8')
                    digits = extract_digits(code)
                    if len(digits) >= 15:
                        return digits, page_num + 1

        pdf.close()
        return None, None
    except Exception as e:
        return None, None

@app.route('/api/process', methods=['POST'])
def process():
    if 'excel' not in request.files or 'pdfs' not in request.files:
        return jsonify({'error': 'Missing files'}), 400

    excel_file = request.files['excel']
    pdf_files = request.files.getlist('pdfs')

    try:
        # Parse Excel
        token_map = parse_excel(excel_file)
        if not token_map:
            return jsonify({'error': 'No valid token/receipt pairs in Excel'}), 400

        results = []
        matched = 0
        unmatched = 0

        for pdf_file in pdf_files:
            pdf_bytes = pdf_file.read()
            filename = pdf_file.filename

            # Decode barcode
            token, page = decode_barcode(pdf_bytes)

            row = {
                'original': filename,
                'token': token or '',
                'newName': '',
                'status': '',
                'reason': ''
            }

            if not token:
                row['status'] = 'unmatched'
                row['reason'] = 'No barcode found'
                unmatched += 1
            else:
                # Match against Excel
                receipt = token_map.get(token)
                if not receipt:
                    row['status'] = 'unmatched'
                    row['reason'] = 'Barcode not in Excel'
                    unmatched += 1
                else:
                    new_name = receipt + '.pdf'
                    row['newName'] = new_name
                    row['status'] = 'matched'
                    matched += 1
                    if page and page > 1:
                        row['reason'] = 'Page ' + str(page)

            results.append(row)

        return jsonify({
            'success': True,
            'results': results,
            'matched': matched,
            'unmatched': unmatched,
            'total': len(results)
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500
