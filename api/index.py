from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
from openpyxl import load_workbook
from pdf2image import convert_from_bytes
from pyzbar.pyzbar import decode
from PIL import Image
import io
import zipfile
import re

app = Flask(__name__)
CORS(app)

def extract_digits(text):
    if text is None:
        return ""
    return re.sub(r'[^0-9]', '', str(text))

def parse_excel(excel_file):
    wb = load_workbook(excel_file)
    token_map = {}
    dup_tokens = set()

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        col_stats = {}
        for row in rows:
            for col_idx, cell in enumerate(row):
                if col_idx not in col_stats:
                    col_stats[col_idx] = {'t20': 0, 't15': 0}
                d = extract_digits(cell)
                if len(d) == 20:
                    col_stats[col_idx]['t20'] += 1
                if len(d) == 15:
                    col_stats[col_idx]['t15'] += 1

        token_col = -1
        for k in col_stats:
            if col_stats[k]['t20'] > (col_stats.get(token_col, {}).get('t20', 0) if token_col >= 0 else 0):
                token_col = k

        target_col = -1
        for k in col_stats:
            if k != token_col and col_stats[k]['t15'] > (col_stats.get(target_col, {}).get('t15', 0) if target_col >= 0 else 0):
                target_col = k

        for row in rows:
            token = ""
            if token_col >= 0 and token_col < len(row):
                token = extract_digits(row[token_col])

            if not token or len(token) != 20:
                continue

            target = ""
            if target_col >= 0 and target_col < len(row):
                target = extract_digits(row[target_col])

            if not target or len(target) != 15:
                continue

            if token in token_map:
                dup_tokens.add(token)
            else:
                token_map[token] = target

    return token_map, dup_tokens

def decode_pdf_barcode(pdf_bytes):
    try:
        images = convert_from_bytes(pdf_bytes, first_page=1, last_page=3)
        for page_num, image in enumerate(images, 1):
            for scale in [1, 0.75, 0.5]:
                img = image.resize((int(image.width * scale), int(image.height * scale))) if scale != 1 else image
                decoded = decode(img)
                if decoded:
                    for obj in decoded:
                        code = obj.data.decode('utf-8')
                        digits = extract_digits(code)
                        if len(digits) >= 15:
                            return digits, page_num
        return None, None
    except:
        return None, None

@app.route('/')
def index():
    with open('public/index.html') as f:
        return f.read()

@app.route('/api/process', methods=['POST'])
def process():
    if 'excel' not in request.files or 'pdfs' not in request.files:
        return jsonify({'error': 'Missing files'}), 400

    excel_file = request.files['excel']
    pdf_files = request.files.getlist('pdfs')

    try:
        token_map, dup_tokens = parse_excel(excel_file)
        if not token_map:
            return jsonify({'error': 'No valid pairs found'}), 400

        results = []
        matched = 0
        unmatched = 0
        used_names = {}

        for pdf_file in pdf_files:
            pdf_bytes = pdf_file.read()
            filename = pdf_file.filename

            token, page = decode_pdf_barcode(pdf_bytes)
            row = {'original': filename, 'token': token or '', 'newName': '', 'status': '', 'reason': ''}

            if not token:
                row['status'] = 'unmatched'
                row['reason'] = 'No barcode found'
                unmatched += 1
            else:
                hit = token_map.get(token)
                if not hit:
                    row['status'] = 'unmatched'
                    row['reason'] = 'Not in Excel'
                    unmatched += 1
                else:
                    n = used_names.get(hit, 0) + 1
                    used_names[hit] = n
                    final_name = hit + '.pdf' if n == 1 else hit + ' (' + str(n) + ').pdf'
                    row['newName'] = final_name
                    row['status'] = 'matched'
                    matched += 1
                    if page and page > 1:
                        row['reason'] = 'Page ' + str(page)

            results.append(row)

        return jsonify({'success': True, 'results': results, 'matched': matched, 'unmatched': unmatched, 'total': len(results)})
    except Exception as e:
        return jsonify({'error': str(e)}), 500
